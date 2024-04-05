# -*- coding: utf-8 -*-
# @Author: Dastan_Alam
# @Date:   24-03-2024 11:31:34 PM       23:31:34
# @Last Modified by:   Dastan_Alam
# @Last Modified time: 25-03-2024 11:50:11 PM       23:50:11
import openpyxl
import requests
from bs4 import BeautifulSoup
import re
import html2text


print("working please wait for 2 min")
print("............")
workbook = openpyxl.Workbook()
sheet = workbook.active


file_path = "coupons.txt"  # Replace "coupons.txt" with the path to your text file

# Open the file in read mode
with open(file_path, "r") as file:
    # Read lines from the file and strip any leading/trailing whitespace
    coup = [line.strip() for line in file.readlines()]

# Now, 'coup' contains the coupon URLs from the file as a list
# print(coup)

def url(coup):
    for index, coupon in enumerate(coup):
        response = requests.get(coupon)
        soup = BeautifulSoup(response.content, 'html.parser')

        # Extracting data
        heading_elem = soup.find('h1', class_='clp-lead__title')
        heading = heading_elem.text.strip() if heading_elem else 'Heading not found'
        
        sub_heading_elem = soup.find('div', class_='ud-text-md clp-lead__headline')
        sub_heading = sub_heading_elem.text.strip() if sub_heading_elem else 'Sub heading not found'
        
        what_youll_learn_elem = soup.find('ul', class_='ud-unstyled-list ud-block-list what-you-will-learn--objectives-list--qsvE2 what-you-will-learn--objectives-list-two-column-layout--ED4as')
        what_youll_learn_section = what_youll_learn_elem.text.strip() if what_youll_learn_elem else 'What you\'ll learn section not found'
        what_youll_learn = "<ul>" + what_youll_learn_section.replace(".", ".\n</ul><ul>") + "</ul>"
        
        requirements_elem = soup.find('ul', class_='ud-unstyled-list ud-block-list')
        requirements_section = requirements_elem.text.strip() if requirements_elem else 'Requirements not found'
        requirements_split = "<ul>" + re.sub(r'(?=[A-Z])', '\n'+"</ul><ul>", requirements_section) + "</ul>"
        
        description_elem = soup.find('div', class_='show-more-module--container--teP7C')
        plain_text = html2text.html2text(description_elem.prettify())
        beautified_text = '\n'.join(line.strip() for line in plain_text.split('\n') if line.strip())
        
        coupon_string = coupon
        
        intro_asset = soup.find(class_='intro-asset--img-aspect--3gluH')
        
        # Find the img tag within the intro_asset
        if intro_asset:
            img_tag = intro_asset.find('img')
            if img_tag:
                # Extract srcset attribute
                srcset = img_tag.get('srcset')
                if srcset:
                    # Split srcset by comma to separate URLs and resolutions
                    sources = srcset.split(',')
                    # Iterate over sources to find the highest resolution image URL
                    highest_resolution_url = None
                    highest_resolution = 0
                    for source in sources:
                        url, resolution = source.strip().split(' ')
                        resolution = int(resolution[:-1])  # remove 'w' at the end and convert to int
                        if resolution > highest_resolution:
                            highest_resolution_url = url
                            highest_resolution = resolution
                    # print("Highest resolution image URL:", highest_resolution_url)
                else:
                    print("No srcset attribute found.")
            else:
                print("No img tag found within the specified class.")
        else:
            print("No element found with class 'intro-asset--img-aspect--3gluH'.")
                
        # Writing to Excel sheet
        sheet.cell(row=index+1, column=1).value = heading
        sheet.cell(row=index+1, column=2).value = f"{sub_heading}\n\n<img src='{highest_resolution_url}'><h2>What you'll learn</h2>{what_youll_learn}\n\n<h2>Requirements</h2>{requirements_split}\n\n<h2>Description</h2>{beautified_text}<br><br><br><a href='{coupon_string}' target='_blank'>GET THE COURES</a>"
        
        # sheet.cell(row=index+1, column=2).value = f"{sub_heading}\n\n<h2>What you'll learn</h2>{what_youll_learn}\n\n<h2>Requirements</h2>{requirements_split}\n\n<h2>Description</h2>{beautified_text}<br><br><br><a href='{coupon_string}' target='_blank'>GET THE COURES</a>"
        
        


# Calling the function to process the URLs
url(coup)

# Saving the workbook
workbook.save('data.xlsx')
print("done.......")
print("opening chrome")
print("......")
