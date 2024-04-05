# -*- coding: utf-8 -*-
# @Author: Dastan_Alam
# @Date:   10-03-2024 12:56:50 PM       12:56:50
# @Last Modified by:   Dastan_Alam
# @Last Modified time: 24-03-2024 11:48:43 PM       23:48:43

import requests
from bs4 import BeautifulSoup
import re
import html2text
import auto_finder
import httplib2
import openpyxl
from oauth2client.client import flow_from_clientsecrets
from oauth2client.file import Storage
from oauth2client.tools import run_flow
from googleapiclient import discovery

workbook = openpyxl.Workbook()
sheet = workbook.active
def url(coupon):
    print(coupon)
    for y in range(len(coupon)):
        # def ex(coupon[y]):
            # URL of the Udemy course
            # Send a GET request to the URL
            response = requests.get(coupon[y])
            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(response.content, 'html.parser')
            # Extract the heading
            heading_elem = soup.find('h1', class_='clp-lead__title')
            heading = heading_elem.text.strip() if heading_elem else 'Heading not found'
            # Extract the subheading
            sub_heading_elem = soup.find('div', class_='ud-text-md clp-lead__headline')
            sub_heading = sub_heading_elem.text.strip() if sub_heading_elem else 'Sub heading not found'
            # Extract the "What you'll learn" section
            what_youll_learn_elem = soup.find('ul', class_='ud-unstyled-list ud-block-list what-you-will-learn--objectives-list--qsvE2 what-you-will-learn--objectives-list-two-column-layout--ED4as')
            what_youll_learn_section = what_youll_learn_elem.text.strip() if what_youll_learn_elem else 'What you\'ll learn section not found'
            # String containing what you'll learn
            what_youll_learn_string = what_youll_learn_section
            what_youll_learn = "<ul>"+what_youll_learn_string.replace(".", ".\n</ul><ul>")
            # Convert the string into a list by splitting on the period ('.') character
            # what_youll_learn_list = what_youll_learn_string.split('.')
            # # Extract the course content section
            # course_content_elem = soup.find('div', class_='ud-text-sm')
            # course_content_section = course_content_elem.text.strip() if course_content_elem else 'Course content not found'
            # course_content_string = course_content_section
            # course_content_list = re.split(r'(?=[A-Z])', course_content_string) 
            # Extract the requirements section
            requirements_elem = soup.find('ul', class_='ud-unstyled-list ud-block-list')
            
            requirements_section = requirements_elem.text.strip() if requirements_elem else 'Requirements not found'
            
            requirements_string = requirements_section
            
            # Split the string using regular expression and insert newline character
            
            requirements_split = "<ul>"+re.sub(r'(?=[A-Z])', '\n'+"</ul><ul>", requirements_string)
            
            # Extract the description section
            
            description_elem = soup.find('div', class_='show-more-module--container--teP7C')
            
            # Print the HTML content of the description
            # print(description_elem.prettify())
        
            # normal print html to text
            # html_code =description_elem.prettify()
            # plain_text = html2text.html2text(html_code)
            # print(plain_text)
            def beautify_text(text):
            
                # Replace consecutive newline characters with just one newline character
            
                return '\n'"<br>".join(line.strip() for line in text.split('\n') if line.strip())
            
            plain_text = html2text.html2text(description_elem.prettify())
            beautified_text = beautify_text(plain_text)
            
            # print(beautified_text)
        
            # Print the extracted content without merging lines
            # print(f"Heading: {heading}")
            # print(" ")
            # print(f"Sub Heading: {sub_heading}")
            # print(" ")
            # print(f"What you'll learn:")
            # Print each item in the list on a separate line
            # for item in what_youll_learn_list:
            # print(item)
            # print(" ")
            # print(f"Course content: {course_content_section}")
            # print(" ")
            # print(f"Requirements: {requirements_list}")
            # # Print each item in the list on a separate line
            # # for item in requirements_list:
            #print(item.strip())  
            # # Stripping whitespace from each item
            # print(" ")
            
            # print(f"Description: {description_list}")
            # Print each item in the list on a separate line
            # print("Description:",beautified_text)
            # print(" ")
            
            coupon_string = ''.join(coupon[y])
            all=sub_heading+"<h2>What you'll learn</h2>"+"\n"+what_youll_learn+"</ul>"+"\n"+"<h2>Requirements</h2>"+requirements_split+"</ul>"+"<h2>Description</h2>"+beautified_text+"<br><br><br>"+coupon_string
            
            sheet.cell(row=y+2, column=1).value = heading
            sheet.cell(row=y+2, column=2).value = all
            
            # print(all)
            
workbook.save('data.xlsx')            
                    
            
        